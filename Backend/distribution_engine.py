"""
Distribution engine
====================
Executes a distribution rule: pull rows from the selected report tables for
the schedule's time window, render each into the chosen format(s)
(PDF / XLSX / CSV), and deliver via email and/or disk.

Data sources:
  * PostgreSQL (SQLAlchemy) — daily/weekly/monthly/detailed/material reports.
  * SQL Server (read-only)  — batch materials.
All rule/state writes go to PostgreSQL only.
"""

import io
import csv
import os
import logging
from datetime import datetime, timedelta, time
from html import escape

from config import SQLSERVER_BATCH_MATERIALS_TABLE, SQLSERVER_DATABASE
from models import db
from models.distribution import DistributionRule
from models.kpi_material import KPIMaterial
import smtp_config
from utils.timezone import format_saudi_display

logger = logging.getLogger(__name__)

DEFAULT_SAVE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'distribution_output')


# ── Report catalog ────────────────────────────────────────────────────────────
# key -> metadata. ``source`` is informational for the UI.
REPORT_CATALOG = [
    {'key': 'daily', 'label': 'Daily Reports', 'description': 'Per-product production summary (SQL Server, read-only)', 'source': 'sqlserver'},
    {'key': 'weekly', 'label': 'Weekly Reports', 'description': 'Per-product production summary (SQL Server, read-only)', 'source': 'sqlserver'},
    {'key': 'monthly', 'label': 'Monthly Reports', 'description': 'Per-product production summary (SQL Server, read-only)', 'source': 'sqlserver'},
    {'key': 'detailed', 'label': 'Detailed Report', 'description': 'Per-batch material detail (SQL Server, read-only)', 'source': 'sqlserver'},
    {'key': 'material', 'label': 'Material Consumption', 'description': 'Planned vs actual material usage (SQL Server, read-only)', 'source': 'sqlserver'},
    {'key': 'batch_historical', 'label': 'Batch Historical Reports', 'description': 'Per-batch material detail (SQL Server, read-only)', 'source': 'sqlserver'},
    {'key': 'batch_raw', 'label': 'Batch Raw Data', 'description': 'Raw batch material rows (SQL Server, read-only)', 'source': 'sqlserver'},
]

VALID_SOURCE_KEYS = {c['key'] for c in REPORT_CATALOG}


def get_report_catalog():
    return REPORT_CATALOG


def _resolve_window(rule):
    """Compute the (from_dt, to_dt) data window for a rule.

    Two modes:
      * ``custom`` — uses the rule's explicit ``custom_start`` / ``custom_end``.
      * ``auto``   — period_back: ``to_dt`` is the most recent occurrence of
        ``window_end_time``; ``from_dt`` = ``to_dt`` minus the period
        (daily 1d / weekly 7d / monthly 30d), placed at ``window_start_time``.
        So daily 07:00→07:00 = "yesterday 07:00 → today 07:00".
    """
    window_mode = getattr(rule, 'window_mode', 'auto') or 'auto'

    if window_mode == 'custom':
        if rule.custom_start and rule.custom_end:
            return rule.custom_start, rule.custom_end
        # Fall back to auto if custom dates are missing.

    now = datetime.now()
    end_time = getattr(rule, 'window_end_time', None) or time(7, 0)
    start_time = getattr(rule, 'window_start_time', None) or time(7, 0)

    # to_dt = most recent past occurrence of end_time
    to_dt = now.replace(hour=end_time.hour, minute=end_time.minute, second=0, microsecond=0)
    if now < to_dt:
        to_dt = to_dt - timedelta(days=1)

    days = {'daily': 1, 'weekly': 7, 'monthly': 30}.get(rule.schedule_type, 1)
    from_day = (to_dt - timedelta(days=days)).date()
    from_dt = datetime.combine(from_day, start_time)
    return from_dt, to_dt


# ── Data fetchers — return (columns, rows) where rows is list[dict] ────────────

def _fetch_batch_materials(from_dt, to_dt, limit=2000):
    """Read-only query against SQL Server batch materials."""
    from sqlalchemy import text
    tbl = SQLSERVER_BATCH_MATERIALS_TABLE
    sql = text(
        f"""
        SELECT
            [Source Server], [OrderID], [Batch Name], [Product Name],
            [Batch Act Start], [Batch Act End], [Quantity],
            [Material Name], [Material Code], [SetPoint Float], [Actual Value Float]
        FROM [{SQLSERVER_DATABASE}].[dbo].[{tbl}]
        WHERE [Batch Act Start] >= :from_dt AND [Batch Act Start] < :to_dt
        ORDER BY [Batch Act Start] DESC
        OFFSET 0 ROWS FETCH NEXT :lim ROWS ONLY
        """
    )
    engine = db.get_engine(bind='sqlserver')
    with engine.connect() as conn:
        result = conn.execute(sql, {'from_dt': from_dt, 'to_dt': to_dt, 'lim': limit})
        columns = list(result.keys())
        raw = result.fetchall()
    rows = []
    for r in raw:
        d = {}
        for i, col in enumerate(columns):
            val = r[i]
            d[col] = val.isoformat() if hasattr(val, 'isoformat') else val
        rows.append(d)
    return columns, rows


def _fetch_batch_rows(from_dt, to_dt, limit=5000):
    """Read-only SQL Server BatchMaterials rows for the window (ascending by start)."""
    return (
        KPIMaterial.query
        .filter(KPIMaterial.batch_act_start >= from_dt,
                KPIMaterial.batch_act_start <= to_dt)
        .order_by(KPIMaterial.batch_act_start.asc())
        .limit(limit)
        .all()
    )


def _aggregate_by_product(from_dt, to_dt):
    """Group BatchMaterials by product (mirrors frontend ``aggregateByProduct``).

    Columns: Product Name | No Of Batches (unique batches) | Sum SP | Sum Act |
    Err Kg | Err %.  ``Err %`` is left numeric-as-string so the PDF can colour it.
    """
    mats = _fetch_batch_rows(from_dt, to_dt)
    groups = {}
    order = []
    for m in mats:
        name = m.product_name or 'Unknown'
        if name not in groups:
            groups[name] = {'sumSP': 0.0, 'sumAct': 0.0, 'batches': set()}
            order.append(name)
        groups[name]['batches'].add(str(m.batch_guid))
        groups[name]['sumSP'] += float(m.setpoint_float or 0)
        groups[name]['sumAct'] += float(m.actual_value_float or 0)

    columns = ['Product Name', 'No Of Batches', 'Sum SP', 'Sum Act', 'Err Kg', 'Err %']
    rows = []
    for name in order:
        g = groups[name]
        err_kg = abs(g['sumAct'] - g['sumSP'])
        err_pct = (err_kg / g['sumSP'] * 100) if g['sumSP'] else 0.0
        rows.append({
            'Product Name': name,
            'No Of Batches': len(g['batches']),
            'Sum SP': f"{g['sumSP']:.2f}",
            'Sum Act': f"{g['sumAct']:.2f}",
            'Err Kg': f"{err_kg:.2f}",
            'Err %': f"{err_pct:.2f}",
        })
    return columns, rows


def _aggregate_by_material(from_dt, to_dt):
    """Group BatchMaterials by material (mirrors frontend ``aggregateByMaterial``).

    Columns: Material Name | Code | Planned (kg) | Actual (kg) | Difference %.
    """
    mats = _fetch_batch_rows(from_dt, to_dt)
    groups = {}
    order = []
    for m in mats:
        name = m.material_name or 'Unknown'
        if name not in groups:
            groups[name] = {'code': m.material_code or '', 'planned': 0.0, 'actual': 0.0}
            order.append(name)
        groups[name]['planned'] += float(m.setpoint_float or 0)
        groups[name]['actual'] += float(m.actual_value_float or 0)

    columns = ['Material Name', 'Code', 'Planned (kg)', 'Actual (kg)', 'Difference %']
    rows = []
    for name in order:
        g = groups[name]
        diff_pct = (abs((g['actual'] - g['planned']) / g['planned']) * 100) if g['planned'] else 0.0
        rows.append({
            'Material Name': name,
            'Code': g['code'],
            'Planned (kg)': f"{g['planned']:.2f}",
            'Actual (kg)': f"{g['actual']:.2f}",
            'Difference %': f"{diff_pct:.2f}",
        })
    return columns, rows


def _fetch_batch_detailed(from_dt, to_dt, limit=5000):
    """Read SQL Server batch materials for the window, grouped by batch.

    Mirrors the frontend "Detailed Report" aggregation
    (``BatchHistoricalReports.tsx`` → ``detailedBatchGroups``):
      * group rows by ``Batch GUID``
      * per material: ``errKg = |actual − setpoint|`` and
        ``errPercent = |errKg / setpoint| * 100``
      * append a per-batch "Total" row summing setpoint/actual.

    Returns a list of groups, each::

        {
          'batchName', 'productName', 'batchStart', 'batchEnd', 'batchQuantity',
          'materials': [{materialName, materialCode, setPoint, actual, errKg, errPercent}, ...],
          'total':     {setPoint, actual, errKg, errPercent},
        }
    """
    mats = _fetch_batch_rows(from_dt, to_dt, limit)

    groups_map = {}
    order = []
    for m in mats:
        key = str(m.batch_guid)
        if key not in groups_map:
            groups_map[key] = {
                'batchName': m.batch_name,
                'productName': m.product_name,
                'batchStart': m.batch_act_start,
                'batchEnd': m.batch_act_end,
                'batchQuantity': m.quantity,
                'materials': [],
            }
            order.append(key)

        sp = m.setpoint_float
        act = m.actual_value_float
        err_kg = abs((act or 0) - (sp or 0)) if (act is not None and sp is not None) else 0.0
        err_pct = abs((err_kg / sp) * 100) if sp else 0.0
        groups_map[key]['materials'].append({
            'materialName': m.material_name,
            'materialCode': m.material_code,
            'setPoint': sp or 0.0,
            'actual': act or 0.0,
            'errKg': err_kg,
            'errPercent': err_pct,
        })

    groups = []
    for key in order:
        g = groups_map[key]
        tot_sp = sum(x['setPoint'] for x in g['materials'])
        tot_act = sum(x['actual'] for x in g['materials'])
        tot_err = abs(tot_act - tot_sp)
        tot_pct = abs((tot_err / tot_sp) * 100) if tot_sp else 0.0
        g['total'] = {'setPoint': tot_sp, 'actual': tot_act, 'errKg': tot_err, 'errPercent': tot_pct}
        groups.append(g)
    return groups


def _fmt_dt(value):
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d %H:%M')
    return str(value)


def _fmt_dt_print(value):
    """Format naive UTC DB datetime for Saudi display in PDF/email."""
    if value is None:
        return 'N/A'
    if not hasattr(value, 'strftime'):
        return str(value)
    return format_saudi_display(value)


# Flat columns used when a batch-detailed report is exported as CSV / XLSX.
DETAILED_FLAT_COLUMNS = [
    'Batch', 'Product', 'Batch Start', 'Batch End', 'Quantity',
    'Material Name', 'Material Code', 'SetPoint', 'Actual', 'Err Kg', 'Err %',
]


def _detailed_to_flat(groups):
    """Flatten batch groups into plain rows for CSV / XLSX (one Total row per batch)."""
    rows = []
    for g in groups:
        for mat in g['materials']:
            rows.append({
                'Batch': g['batchName'] or '',
                'Product': g['productName'] or '',
                'Batch Start': _fmt_dt(g['batchStart']),
                'Batch End': _fmt_dt(g['batchEnd']),
                'Quantity': g['batchQuantity'] if g['batchQuantity'] is not None else '',
                'Material Name': mat['materialName'] or '',
                'Material Code': mat['materialCode'] or '',
                'SetPoint': f"{mat['setPoint']:.2f}",
                'Actual': f"{mat['actual']:.2f}",
                'Err Kg': f"{mat['errKg']:.2f}",
                'Err %': f"{mat['errPercent']:.2f}",
            })
        t = g['total']
        rows.append({
            'Batch': g['batchName'] or '',
            'Product': g['productName'] or '',
            'Batch Start': _fmt_dt(g['batchStart']),
            'Batch End': _fmt_dt(g['batchEnd']),
            'Quantity': g['batchQuantity'] if g['batchQuantity'] is not None else '',
            'Material Name': 'Total',
            'Material Code': '',
            'SetPoint': f"{t['setPoint']:.2f}",
            'Actual': f"{t['actual']:.2f}",
            'Err Kg': f"{t['errKg']:.2f}",
            'Err %': f"{t['errPercent']:.2f}",
        })
    return DETAILED_FLAT_COLUMNS, rows


def _fetch_source(source_key, from_dt, to_dt):
    """Dispatch a catalog key to its data fetcher.

    Returns a dataset dict::

        {'kind': 'flat'|'detailed', 'columns': [...], 'rows': [...], 'groups': [...]|None}
    """
    if source_key in ('daily', 'weekly', 'monthly'):
        # All three are product summaries over the rule's resolved window
        # (daily=1d / weekly=7d / monthly=30d), mirroring the frontend.
        cols, rows = _aggregate_by_product(from_dt, to_dt)
        return {'kind': 'flat', 'columns': cols, 'rows': rows, 'groups': None}
    if source_key == 'material':
        cols, rows = _aggregate_by_material(from_dt, to_dt)
        return {'kind': 'flat', 'columns': cols, 'rows': rows, 'groups': None}
    if source_key in ('detailed', 'batch_historical'):
        groups = _fetch_batch_detailed(from_dt, to_dt)
        cols, rows = _detailed_to_flat(groups)
        return {'kind': 'detailed', 'columns': cols, 'rows': rows, 'groups': groups}
    if source_key == 'batch_raw':
        cols, rows = _fetch_batch_materials(from_dt, to_dt)
        return {'kind': 'flat', 'columns': cols, 'rows': rows, 'groups': None}
    raise ValueError(f'Unknown report source: {source_key}')


def _label_for(source_key):
    for c in REPORT_CATALOG:
        if c['key'] == source_key:
            return c['label']
    return source_key


# ── Renderers — each returns bytes ─────────────────────────────────────────────

def _render_csv(columns, rows):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction='ignore')
    writer.writeheader()
    for row in rows:
        writer.writerow({c: row.get(c, '') for c in columns})
    return buf.getvalue().encode('utf-8-sig')


def _render_xlsx(title, columns, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = (title or 'Report')[:31]

    header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, min(40, len(str(col)) + 4))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── PDF rendering (ReportLab) — mirrors BatchHistoricalReports print layout ────

# Color thresholds for the ERR% column — < 5% is "good" (green), else red.
# Matches the frontend's `error-positive` / `error-negative` classes.
_TEAL = '#0088a9'
_ERR_OK_COLOR = '#28a745'
_ERR_BAD_COLOR = '#dc3545'
_BATCH_BG = '#f7fbfd'
_TOTAL_BG = '#e8f4f8'
_FILTER_BG = '#f3fafc'
_FILTER_BORDER = '#d6eef4'
_GRID = '#dddddd'

_LOGO_CACHE = {}


def _logo_png_bytes(filename):
    """Return PNG bytes for a logo in the frontend assets folder.

    WebP is rasterised to PNG via Pillow. Result is cached; failures degrade
    to ``None`` (no image).
    """
    if filename in _LOGO_CACHE:
        return _LOGO_CACHE[filename]

    base = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(base, '..', 'frontend-package', 'client', 'src', 'assets', filename)
    data = None
    try:
        ext = os.path.splitext(filename)[1].lower()
        if ext == '.webp':
            from PIL import Image
            img = Image.open(path).convert('RGBA')
            buf = io.BytesIO()
            img.save(buf, format='PNG')
            data = buf.getvalue()
        else:
            with open(path, 'rb') as f:
                data = f.read()
    except Exception as e:  # pragma: no cover - logo is best-effort
        logger.warning('Could not load logo %s: %s', filename, e)
    _LOGO_CACHE[filename] = data
    return data


def _logo_flowable(filename, max_h=58, max_w=190):
    """Build an aspect-correct ReportLab Image for a logo, or '' if unavailable."""
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.utils import ImageReader
    data = _logo_png_bytes(filename)
    if not data:
        return ''
    try:
        iw, ih = ImageReader(io.BytesIO(data)).getSize()
        scale = min(max_w / iw, max_h / ih)
        return RLImage(io.BytesIO(data), width=iw * scale, height=ih * scale)
    except Exception as e:  # pragma: no cover
        logger.warning('Could not size logo %s: %s', filename, e)
        return ''


def _display_title(title):
    """Append 'Report' only if the label doesn't already end with it."""
    t = str(title or '').strip()
    low = t.lower()
    if low.endswith('report') or low.endswith('reports'):
        return t
    return f'{t} Report'


def _pdf_styles():
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from reportlab.lib import colors

    base = ParagraphStyle('cell', fontName='Helvetica', fontSize=7.5, leading=9.5,
                           textColor=colors.HexColor('#111111'))
    return {
        'title': ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=17, leading=20,
                                 textColor=colors.HexColor(_TEAL)),
        'sub': ParagraphStyle('sub', fontName='Helvetica', fontSize=8.5, leading=11,
                              textColor=colors.HexColor('#555555')),
        'filt_head': ParagraphStyle('filt_head', fontName='Helvetica-Bold', fontSize=10, leading=13,
                                    textColor=colors.HexColor(_TEAL)),
        'filt_item': ParagraphStyle('filt_item', fontName='Helvetica', fontSize=8.5, leading=11,
                                    textColor=colors.HexColor('#333333')),
        'th': ParagraphStyle('th', fontName='Helvetica-Bold', fontSize=7.5, leading=9.5,
                             textColor=colors.white),
        'cell': base,
        'cell_b': ParagraphStyle('cell_b', parent=base, fontName='Helvetica-Bold'),
        'num': ParagraphStyle('num', parent=base, alignment=TA_RIGHT),
        'num_b': ParagraphStyle('num_b', parent=base, fontName='Helvetica-Bold', alignment=TA_RIGHT),
        'batch': ParagraphStyle('batch', fontName='Helvetica', fontSize=7.5, leading=10.5,
                                textColor=colors.HexColor('#222222')),
        'err_ok': ParagraphStyle('err_ok', parent=base, fontName='Helvetica-Bold', alignment=TA_RIGHT,
                                 textColor=colors.HexColor(_ERR_OK_COLOR)),
        'err_bad': ParagraphStyle('err_bad', parent=base, fontName='Helvetica-Bold', alignment=TA_RIGHT,
                                  textColor=colors.HexColor(_ERR_BAD_COLOR)),
        'footer': ParagraphStyle('footer', fontName='Helvetica', fontSize=7.5, leading=10,
                                 alignment=TA_CENTER, textColor=colors.HexColor('#666666')),
    }


def _pdf_header_story(title, window, total_rows, usable_w, styles):
    """Logo band + title + Report Filters box — returned as a list of flowables."""
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors

    # Right group: Fakieh then ASM, sitting together (right-aligned within their cell).
    right_inner = Table([[_logo_flowable('fakiehlogo.webp', max_h=50, max_w=120),
                          _logo_flowable('Asm_Logo.png', max_h=50, max_w=120)]])
    right_inner.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    # Hercules on the left, Fakieh + ASM on the right.
    logo_tbl = Table([[_logo_flowable('Hercules_New.png', max_h=58, max_w=190), right_inner]],
                     colWidths=[usable_w * 0.42, usable_w * 0.58])
    logo_tbl.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 1.5, colors.HexColor(_TEAL)),
    ]))

    date_range = f"{_fmt_dt(window[0])} to {_fmt_dt(window[1])}" if window else '—'
    filt = Table([
        [Paragraph('Report Filters', styles['filt_head'])],
        [Paragraph(f'<b>Date Range:</b> {escape(date_range)}', styles['filt_item'])],
        [Paragraph(f'<b>Total Records:</b> {total_rows}', styles['filt_item'])],
    ], colWidths=[usable_w])
    filt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(_FILTER_BG)),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor(_FILTER_BORDER)),
        ('LINEBELOW', (0, 0), (0, 0), 0.4, colors.HexColor(_FILTER_BORDER)),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    return [
        logo_tbl, Spacer(1, 10),
        Paragraph(_display_title(title), styles['title']),
        Paragraph('Generated on: ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S'), styles['sub']),
        Spacer(1, 8), filt, Spacer(1, 12),
    ]


def _err_para(value, styles):
    """Colored ERR% Paragraph: green when < 5%, red otherwise."""
    from reportlab.platypus import Paragraph
    try:
        num = float(value)
    except (TypeError, ValueError):
        num = 0.0
    st = styles['err_ok'] if num < 5 else styles['err_bad']
    return Paragraph(f'{num:.2f}', st)


def _usable_width():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    return A4[0] - 2 * 1.2 * cm


def _build_pdf(story):
    from reportlab.platypus import SimpleDocTemplate
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                            topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    doc.build(story)
    return out.getvalue()


def _render_pdf_detailed(title, groups, window):
    """Per-batch PDF with a left rowspan batch column, Total rows and colored ERR%.

    Mirrors the frontend Detailed Report print layout. Uses ReportLab so column
    widths and the merged batch cell render exactly. Falls back to a flat table
    if a batch block can't be laid out (e.g. a single batch taller than a page).
    """
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors

    styles = _pdf_styles()
    usable = _usable_width()
    fracs = [0.17, 0.22, 0.10, 0.13, 0.13, 0.12, 0.13]
    col_w = [usable * f for f in fracs]

    headers = ['Batch', 'Material Name', 'Code', 'Set Point', 'Actual', 'Err Kg', 'Err %']
    data = [[Paragraph(h, styles['th']) for h in headers]]
    cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(_TEAL)),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(_GRID)),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]

    total_rows = 0
    r = 1
    for g in groups:
        mats = g['materials']
        batch_para = Paragraph(
            f'<font color="{_TEAL}"><b>Batch:</b></font> {escape(str(g["batchName"] or "N/A"))}<br/>'
            f'<font color="{_TEAL}"><b>Product:</b></font> {escape(str(g["productName"] or "N/A"))}<br/>'
            f'<font color="{_TEAL}"><b>Start:</b></font> {escape(_fmt_dt_print(g["batchStart"]))}<br/>'
            f'<font color="{_TEAL}"><b>End:</b></font> {escape(_fmt_dt_print(g["batchEnd"]))}<br/>'
            f'<font color="{_TEAL}"><b>Quantity:</b></font> '
            f'{escape(str(g["batchQuantity"]) if g["batchQuantity"] is not None else "N/A")}',
            styles['batch'])

        start_r = r
        for i, mat in enumerate(mats):
            data.append([
                batch_para if i == 0 else '',
                Paragraph(escape(str(mat['materialName'] or '')), styles['cell']),
                Paragraph(escape(str(mat['materialCode'] or '')), styles['cell']),
                Paragraph(f"{mat['setPoint']:.2f}", styles['num']),
                Paragraph(f"{mat['actual']:.2f}", styles['num']),
                Paragraph(f"{mat['errKg']:.2f}", styles['num_b']),
                _err_para(mat['errPercent'], styles),
            ])
            r += 1
        if mats:
            cmds.append(('SPAN', (0, start_r), (0, r - 1)))
            cmds.append(('BACKGROUND', (0, start_r), (0, r - 1), colors.HexColor(_BATCH_BG)))

        t = g['total']
        data.append([
            '',
            Paragraph('Total', styles['cell_b']), '',
            Paragraph(f"{t['setPoint']:.2f}", styles['num_b']),
            Paragraph(f"{t['actual']:.2f}", styles['num_b']),
            Paragraph(f"{t['errKg']:.2f}", styles['num_b']),
            Paragraph(f"{t['errPercent']:.2f}", styles['num_b']),
        ])
        cmds.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor(_TOTAL_BG)))
        r += 1
        total_rows += len(mats) + 1

    table = Table(data, colWidths=col_w, repeatRows=1)
    table.setStyle(TableStyle(cmds))

    story = _pdf_header_story(title, window, total_rows, usable, styles)
    story.append(table)
    story.append(Spacer(1, 10))
    story.append(Paragraph(f'Total Records: {total_rows} | Generated by Fakieh Reporting',
                           styles['footer']))
    try:
        return _build_pdf(story)
    except Exception as e:
        logger.warning('Detailed PDF layout failed (%s); falling back to flat table', e)
        cols, rows = _detailed_to_flat(groups)
        return _render_pdf_flat(title, cols, rows, window)


def _render_pdf_flat(title, columns, rows, window):
    """Styled flat-table PDF for non-batch report sources (ReportLab)."""
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors

    styles = _pdf_styles()
    usable = _usable_width()
    ncols = max(1, len(columns))
    col_w = [usable / ncols] * ncols

    err_cols = {'Err %', 'Difference %'}
    data = [[Paragraph(escape(str(c)), styles['th']) for c in columns]]
    for row in rows:
        cells = []
        for c in columns:
            val = str(row.get(c, ''))
            if c in err_cols and val != '':
                cells.append(_err_para(val, styles))  # green < 5%, red otherwise
            else:
                cells.append(Paragraph(escape(val), styles['cell']))
        data.append(cells)

    table = Table(data, colWidths=col_w, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(_TEAL)),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(_GRID)),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    story = _pdf_header_story(title, window, len(rows), usable, styles)
    story.append(table)
    story.append(Spacer(1, 10))
    story.append(Paragraph(f'Total Records: {len(rows)} | Generated by Fakieh Reporting',
                           styles['footer']))
    return _build_pdf(story)


def _render(fmt, title, dataset, window=None):
    columns, rows = dataset['columns'], dataset['rows']
    if fmt == 'csv':
        return _render_csv(columns, rows)
    if fmt == 'xlsx':
        return _render_xlsx(title, columns, rows)
    if fmt == 'pdf':
        if dataset['kind'] == 'detailed' and dataset.get('groups') is not None:
            return _render_pdf_detailed(title, dataset['groups'], window)
        return _render_pdf_flat(title, columns, rows, window)
    raise ValueError(f'Unknown format: {fmt}')


# Inline logo CIDs referenced by the email HTML. The matching image bytes are
# attached as related/inline parts by smtp_config (see _email_inline_logos).
EMAIL_LOGOS = [
    ('hercules_logo', 'Hercules_New.png'),
    ('fakieh_logo', 'fakiehlogo.webp'),
    ('asm_logo', 'Asm_Logo.png'),
]


def _email_inline_logos():
    """Return [(cid, filename, png_bytes), ...] for logos that loaded OK."""
    images = []
    for cid, filename in EMAIL_LOGOS:
        data = _logo_png_bytes(filename)
        if data:
            images.append((cid, filename, data))
    return images


def _fmt_email_dt(dt):
    """Match the dashboard date style, e.g. '10/06/2026, 09:12'."""
    try:
        return dt.strftime('%d/%m/%Y, %H:%M')
    except Exception:
        return str(dt)


def _build_email_html(rule_name, sources, from_dt, to_dt, file_names):
    """Professional dark report card with inline (CID) Hercules / Fakieh / ASM
    logos. Email-safe: table-based layout with inline styles only."""
    title = escape(rule_name or 'Scheduled Report')
    attach_rows = ''.join(
        f'<div style="color:#e5e7eb;font-size:14px;line-height:1.5;'
        f'word-break:break-all">{escape(fn)}</div>'
        for fn in file_names
    ) or '<div style="color:#e5e7eb;font-size:14px">—</div>'
    generated = _fmt_email_dt(datetime.now())
    period_from = _fmt_email_dt(from_dt)
    period_to = _fmt_email_dt(to_dt)

    label_style = ('margin:0 0 4px;font-size:11px;font-weight:700;'
                   'letter-spacing:1.5px;color:#8a93a3;text-transform:uppercase')

    return f"""\
<body style="margin:0;padding:0;background:#3a3d42">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
         style="background:#3a3d42;padding:32px 0">
    <tr>
      <td align="center">
        <table role="presentation" width="600" cellpadding="0" cellspacing="0"
               style="width:600px;max-width:600px;background:#26282d;border-radius:10px;
                      overflow:hidden;font-family:Arial,Helvetica,sans-serif">
          <!-- Header: Hercules left, Fakieh + ASM right -->
          <tr>
            <td style="padding:28px 36px 8px">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                <tr>
                  <td align="left" valign="middle">
                    <img src="cid:hercules_logo" alt="Hercules" height="34"
                         style="display:block;height:34px;width:auto;border:0">
                  </td>
                  <td align="right" valign="middle">
                    <img src="cid:fakieh_logo" alt="Fakieh" height="30"
                         style="display:inline-block;height:30px;width:auto;border:0;
                                vertical-align:middle;margin-right:14px">
                    <img src="cid:asm_logo" alt="ASM" height="30"
                         style="display:inline-block;height:30px;width:auto;border:0;
                                vertical-align:middle">
                  </td>
                </tr>
              </table>
            </td>
          </tr>
          <!-- Title -->
          <tr>
            <td align="center" style="padding:18px 36px 0">
              <p style="{label_style}">Scheduled Report</p>
              <h1 style="margin:6px 0 0;font-size:26px;font-weight:700;color:#ffffff">
                {title}</h1>
            </td>
          </tr>
          <!-- Detail box -->
          <tr>
            <td style="padding:24px 36px 0">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                     style="border:1px solid #3c3f46;border-radius:8px">
                <tr>
                  <td style="padding:16px 20px;border-bottom:1px solid #3c3f46">
                    <p style="{label_style}">Report Period</p>
                    <div style="color:#e5e7eb;font-size:14px">
                      <b style="color:#ffffff">From:</b> {period_from}
                      &nbsp;&mdash;&nbsp;
                      <b style="color:#ffffff">To:</b> {period_to}
                    </div>
                  </td>
                </tr>
                <tr>
                  <td style="padding:16px 20px;border-bottom:1px solid #3c3f46">
                    <p style="{label_style}">Generated</p>
                    <div style="color:#e5e7eb;font-size:14px">{generated}</div>
                  </td>
                </tr>
                <tr>
                  <td style="padding:16px 20px">
                    <p style="{label_style}">Attachment</p>
                    {attach_rows}
                  </td>
                </tr>
              </table>
            </td>
          </tr>
          <!-- Body copy -->
          <tr>
            <td style="padding:22px 36px 4px">
              <p style="margin:0;color:#aab1bd;font-size:14px;line-height:1.6">
                Please find the scheduled report <b style="color:#e5e7eb">{title}</b>
                attached to this email. The attachment uses the same layout as the
                dashboard viewer. The report covers the period shown above. If you
                have questions about this report, please contact your system
                administrator.
              </p>
            </td>
          </tr>
          <!-- Footer -->
          <tr>
            <td style="padding:24px 36px 28px">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                <tr>
                  <td align="center" style="border-top:1px solid #3c3f46;padding-top:18px">
                    <p style="margin:0;color:#6b7280;font-size:12px">
                      This is an automated email from the
                      <b style="color:#9aa3af">Hercules Reporting Module</b>.
                    </p>
                  </td>
                </tr>
              </table>
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>"""


def _save_to_disk(save_path, filename, content_bytes):
    target_dir = save_path.strip() if save_path and save_path.strip() else DEFAULT_SAVE_DIR
    os.makedirs(target_dir, exist_ok=True)
    resolved = os.path.realpath(os.path.join(target_dir, filename))
    with open(resolved, 'wb') as f:
        f.write(content_bytes)
    return resolved


# ── Main entry point ────────────────────────────────────────────────────────

def execute_distribution_rule(rule_id):
    """Execute one rule. Returns {'success': bool, 'message'|'error': str}."""
    rule = DistributionRule.query.get(rule_id)
    if not rule:
        return {'success': False, 'error': f'Rule {rule_id} not found'}

    sources = [s for s in (rule.report_sources or []) if s in VALID_SOURCE_KEYS]
    formats = [f for f in (rule.formats or []) if f in ('pdf', 'xlsx', 'csv')]
    if not sources:
        return {'success': False, 'error': 'No valid report sources configured'}
    if not formats:
        return {'success': False, 'error': 'No output formats selected'}

    try:
        from_dt, to_dt = _resolve_window(rule)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        attachments = []  # (filename, bytes)
        total_rows = 0

        for source_key in sources:
            dataset = _fetch_source(source_key, from_dt, to_dt)
            total_rows += len(dataset['rows'])
            title = _label_for(source_key)
            for fmt in formats:
                content = _render(fmt, title, dataset, window=(from_dt, to_dt))
                filename = f'{source_key}_{timestamp}.{fmt}'
                attachments.append((filename, content))

        file_names = [fn for fn, _ in attachments]
        delivery = rule.delivery_method or 'email'
        messages = []

        if delivery in ('email', 'both'):
            recipients = rule.recipients or []
            if not recipients:
                raise ValueError('Email delivery selected but no recipients configured')
            subject = f"{rule.name or 'Scheduled Report'} — {datetime.now().strftime('%Y-%m-%d')}"
            body = _build_email_html(rule.name, sources, from_dt, to_dt, file_names)
            res = smtp_config.send_email(
                recipients, subject, body,
                attachments=attachments,
                inline_images=_email_inline_logos(),
            )
            if not res.get('success'):
                raise RuntimeError(res.get('error', 'Email send failed'))
            messages.append(f'emailed {len(recipients)} recipient(s)')

        if delivery in ('disk', 'both'):
            saved = [_save_to_disk(rule.save_path, fn, content) for fn, content in attachments]
            messages.append(f'saved {len(saved)} file(s) to disk')

        rule.last_run_at = datetime.utcnow()
        rule.last_run_status = 'success'
        rule.last_run_error = None
        db.session.commit()
        window = f"{from_dt.strftime('%b %d %H:%M')} → {to_dt.strftime('%b %d %H:%M')}"
        return {
            'success': True,
            'message': f"Delivered {total_rows} row(s) for {window} — " + ', '.join(messages),
        }

    except Exception as e:
        logger.error('Error executing distribution rule %s: %s', rule_id, e, exc_info=True)
        try:
            rule.last_run_at = datetime.utcnow()
            rule.last_run_status = 'error'
            rule.last_run_error = str(e)
            db.session.commit()
        except Exception:
            db.session.rollback()
        return {'success': False, 'error': str(e)}
